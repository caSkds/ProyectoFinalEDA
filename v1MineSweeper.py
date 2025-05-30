import tkinter as tk
from tkinter import messagebox
import random
import sys
import time

from openpyxl import Workbook, load_workbook
import os

# Variables globales del minijuego
secuencia = []#Donde se va a guardar el minigame
flechas = {1: "↑", 2: "↓", 3: "←", 4: "→"}#El diccionario para la secuencia
temporizadorPaTodos = None#Es como un ID para el temporizador en caso de que vuelva a caer; se reinicie
tiempo_inicio = None#Para que se vaya actualizando el tiempo conforme pase el tiempo
minaActual= None#Variable que es usada para almacenar las coordenadas de la bomba que presionó el jugador
#Las variables de mi tio
myApp = tk.Tk()
startTime = time.time()
finalTime = 0 # Calculado al final 

gameFrame = tk.Frame(myApp)
gameFrame.pack()

loseFrame = tk.Frame(myApp,width =300, height =200)
# Esta es una prueba de la pantalla de derrota
#Libertad de modificarla
# Más detalles están en la función lose(), casi hasta abajo del código
loseLabel = tk.Label(loseFrame, text = "You Lose")

# Frame del minijuego
minijuegoFrame = tk.Frame(myApp)

# Variables del minijuego en el frame
muestraFlechas = tk.Label(minijuegoFrame)
info = tk.Label(minijuegoFrame)
etiqueta_tiempo = tk.Label(minijuegoFrame)

#Número de poderes en el tablero
bombs = 10
shields = 5
radars = 3
powerUps = 2

remainingShields = 0

map = []
mapGenerated = False
buttons = []
specialButtons = {
    "shield" : 100,
    "radar" : 200,
    "powerUp" : 300,
    "bombs" : 500
}

# Altura del juego
height  = 15
# Ancho del juego
width = 10
myApp.title("Mine Sweeper")

# Create a height x width grid for the matrix representation of the game 
# Create a height x width grid for the buttons
for i in range(0,height):
    newList = []
    for j in range(0,width):
        newList.append(0)
    map.append(newList)


for i in range(0,height):
    otherNewList = []
    for j in range(0,width):
        otherNewList.append(0)
    buttons.append(otherNewList)

#FUNCIONES DEL MINIJUEGO
def actualizarTemporizador():
    global temporizadorPaTodos, tiempo_inicio #Variables que se van a ir actualizando
    tiempoTranscurrido = int(time.time() - tiempo_inicio) #El tiempo de inicio se va a ir actualizando conforme al tiempo
    tiempoRestante = 4 - tiempoTranscurrido #Aquí podríamos añadir mas tiempo; la verdad no se
    etiqueta_tiempo.config(text=f'Tiempo: {tiempoRestante} s') #Esta el la label para el tiempo
    
    if tiempoRestante > 0:
        temporizadorPaTodos = myApp.after(100, actualizarTemporizador) #Se va actualizando el id del temporizador
    else:
        info.config(text="Se acabó el tiempo")#Si se llegó a 0; se acaba el juego
        muestraFlechas.config(text="") #Desaparecen las flechas
        myApp.unbind('<Key>') #Si se modifica; pongan Key con K no k; perdi mas de 2 horas dx
        # Si se acaba el tiempo, va a la pantalla de derrota
        irADerrota()#Se murio la derrota

def iniciarJuego():
    global secuencia, tiempo_inicio, temporizadorPaTodos
    if temporizadorPaTodos:
        myApp.after_cancel(temporizadorPaTodos) #Si ya habia un temporizador; lo elimina y empieza uno nuevo
    tiempo_inicio = time.time() #El tiempo de inicio inicia en 0
    secuencia = [random.randint(1, 4) for i in range(random.randint(5, 10))] #Se va a ingresar a la secuencia 5 a 10 numeros entre 1 y 4
    muestraFlechas.config(text=' '.join([flechas[num] for num in secuencia])) #Se muestran las flechas en la etiqueta
    info.config(text="¡Ingresa la secuencia!") #La label informativa 
    actualizarTemporizador() #Se inicia el nuevo temporizador
    myApp.bind('<Key>', teclaPresionada) #Se habilitan las flechas

def teclaPresionada(tecla):
    global secuencia
    convertidor = {'Up': 1, 'Down': 2, 'Left': 3, 'Right': 4} #El diccionario convierte las teclas ingresadas en numeros
    if secuencia: #Mientras la secuencia contenga algo:
        if convertidor[tecla.keysym] == secuencia[0]: #Si el jugador le atinó:
            secuencia.pop(0) #Elimina de la secuencia el numero ingresado
            muestraFlechas.config(text=' '.join([flechas[num] for num in secuencia])) #Se actualiza el visor de flechas
            if not secuencia: #Si acabó satisfactoriamente la secuencia:
                info.config(text="Bomba desarmada") #Muestra la leyenda
                myApp.unbind('<Key>') #Se quita lo de las key
                if temporizadorPaTodos:
                    myApp.after_cancel(temporizadorPaTodos) #Detiene y cancela el temporizador
                # Si gana, regresa al juego principal
                volverAlJuego() #Se continua en la partida
        else:
            info.config(text="La bomba explotó") #Si falla; pierde la partida
            muestraFlechas.config(text="") #Ya no se muestra la secuencia en el label
            myApp.unbind('<Key>') #Se quitan los permisos para las flechitas
            if temporizadorPaTodos:
                myApp.after_cancel(temporizadorPaTodos) #Se detiene el temporizador
            # Si explota la bomba, va a la pantalla de derrota
            irADerrota() #Va a la pantalla de derrota

def mostrarMinijuego():
    gameFrame.pack_forget() #Se oculta el widget del buscaminas
    minijuegoFrame.pack()  #Se muestra el minijuego
    muestraFlechas.pack() #Se muestra la secuencia
    info.pack() #Se muestra la info
    etiqueta_tiempo.pack() #Se muestra el tiempo
    iniciarJuego() #Se inicia el juego

def volverAlJuego():
    minijuegoFrame.pack_forget() #Se oculta el minijuego si se gana
    muestraFlechas.config(text="") #Se ocultan las labels
    info.config(text="") 
    etiqueta_tiempo.config(text="")
    gameFrame.pack() #Se vuelve al buscaminas
    global minaActual
    if minaActual is not None:
        h, w = minaActual
        map[h][w] = -1
        buttons[h][w]["text"] = "S"
        mina_actual = None

def irADerrota():
    minijuegoFrame.pack_forget() #Se oculta el minijuego
    lose() #Se pierde


#FUNCIONES DEL BUSCAMINAS
#Función para visualizar el mapa en la terminal
#Solo usar para depuración 
def printMap(gameMap: list):
    for i in range(0,len(gameMap)):
        print(gameMap[i])

# Adding every game object:
def addGameObject(gameMap: list, x_avoid :int, y_avoid : int, nProperty: int, propertyName :str):
    # Making sure the first cell is right next to a bomb
    startingX = x_avoid
    startingY = y_avoid
    if propertyName == "bombs":
        if startingX == width - 1:
            startingX = x_avoid -1
        if startingY == height - 1:
            startingY = y_avoid - 1
        if x_avoid != width - 1 and y_avoid != height - 1:
            gameMap[startingY+1][startingX+1] = specialButtons[propertyName]
        else:
            gameMap[startingY][startingX] = specialButtons[propertyName]
        nProperty -= 1
            
    nProperty -= 1
    while nProperty > 0:
        for i in range(0,height):
            for j in range(0,width):
                if i!= y_avoid and j!=x_avoid and gameMap[i][j] == 0:
                    if random.randint(0,100)<5 and nProperty > 0:
                        gameMap[i][j] = specialButtons[propertyName]
                        nProperty -= 1

# Initialize the map with 0s and place all objects randomly
def initializeMap(gameMap: list, x_0 : int, y_0:int, nBombs: int, nShields: int, nRadars:int, nPowerUps: int):
    global mapGenerated
    addGameObject(gameMap, x_0, y_0, nBombs, "bombs")
    addGameObject(gameMap, x_0, y_0, nShields, "shield")
    addGameObject(gameMap, x_0, y_0, nRadars, "radar")
    addGameObject(gameMap, x_0, y_0, nPowerUps, "powerUp")
    """
    # Fills the map with number of surrounding bombs
    for i in range(0,height):
            for j in range(0,width):
                if map[i][j] != 500:
                    map[i][j] = getNeighborBombs(map,j,i)
    """
    mapGenerated = True




# Gets the count of bombs surrounding a cell
def getNeighborBombs(gameMap: list, x : int, y : int) -> int:
    count = 0
    for i in range(-1, 2):
        for j in range(-1, 2):
            if 0 <= x + i < width and 0 <= y + j < height:
                if gameMap[y + j][x + i] == 500:
                    count += 1
    return count




def radarEffect(gameMap: list, x: int, y: int, buttons: list): 
    # This function will reveal all the cells in a 3x3 area around the radar
    for i in range(-1, 2):
        for j in range(-1, 2):
            if 0 <= x + i < width and 0 <= y + j < height:
                if gameMap[y + j][x + i] == 0:
                    buttons[y + j][x + i]["text"] = getNeighborBombs(gameMap, x + i, y + j)
                else:
                    buttons[y + j][x + i]["text"] = gameMap[y + j][x + i]


def calcularPuntaje():
    final_time = time.time() - startTime
    puntaje = round(final_time, 2)
    return puntaje

def lose():
    global finalTime
    finalTime = time.time() - startTime
    print(finalTime)
    gameFrame.pack_forget()
    loseFrame.pack()
    loseLabel.grid(row = 0, column = 0)
    losetime = tk.Label(loseFrame, text = str(round(finalTime, 2)))
    losetime.grid(row = 0, column = 1)

    # pedir nombre del jugador
    def guardar():
        nombre = nombreEntry.get()
        if nombre:
            puntaje = calcularPuntaje()  
            guardarPuntaje(nombre, puntaje)
            mostrarTop3()

    nombreLabel = tk.Label(loseFrame, text="Nombre:")
    nombreLabel.grid(row=1, column=0)
    nombreEntry = tk.Entry(loseFrame)
    nombreEntry.grid(row=1, column=1)

    guardarBtn = tk.Button(loseFrame, text="Guardar Puntaje", command=guardar)
    guardarBtn.grid(row=2, column=0, columnspan=2)


# Defines behavior for left click on a button
def buttonClick(gameMap: list, name : str,buttonSet : list):
    global mapGenerated
    global remainingShields

    buttonHeight =int(name[:name.index('F')])
    buttonWidth = int(name[name.index('e')+1:])
    if not mapGenerated:
        initializeMap(gameMap, buttonWidth, buttonHeight, bombs, shields, radars, powerUps)
        printMap(gameMap)
    if gameMap[buttonHeight][buttonWidth]==-1: #Siempre que se desactive una bomba con exito; el valor de esa casilla va a valer -1
        return
    if gameMap[buttonHeight][buttonWidth] == 500:
        if remainingShields == 0:
            global minaActual
            minaActual=(buttonHeight,buttonWidth)
            mostrarMinijuego()
        else:
            remainingShields-=1
    elif gameMap[buttonHeight][buttonWidth] == 100:
        buttonSet[buttonHeight][buttonWidth]["text"] = "SHIELD"
        
    elif gameMap[buttonHeight][buttonWidth] == 200:
        radarEffect(gameMap, buttonWidth, buttonHeight, buttonSet)
    elif gameMap[buttonHeight][buttonWidth] == 300: 
        buttonSet[buttonHeight][buttonWidth]["text"] = "POWER UP"
    else:
        buttonSet[buttonHeight][buttonWidth]["text"] =getNeighborBombs(gameMap,buttonWidth, buttonHeight)

# Defines flagging a button
def flag( event ):
    event.widget.config(text="F")




# Create buttons for the grid
for i in range(0,height):
    for j in range(0,width):
        currentName = str(i)+str(False)+str(j)
        button  =tk.Button(gameFrame,
                name = str(i)  +str(j),
                command =  lambda buttonName = currentName: buttonClick(map,buttonName,buttons))
        #print(str(i)+str(j))
        button.grid(
            row = i,
            column = j
        )
        #checks os for right click
        if sys.platform == "darwin":
            button.bind("<Button-2>", flag)
        else:
            button.bind("<Button-3>", flag)
        #print(type(map[i][j]))
        buttons[i][j] = button

def quicksort(lista):
    if len(lista) <= 1:
        return lista
    
    pivote = lista[0]
    iguales = [pivote]
    mayores = []
    menores = []

    for elemento in lista[1:]:
        if elemento[1] > pivote[1]:
            mayores.append(elemento)
        elif elemento[1] < pivote[1]:
            menores.append(elemento)
        else:
            iguales.append(elemento)

    return quicksort(mayores) + iguales + quicksort(menores)

def guardarPuntaje(nombre, puntaje, archivo="puntajes.xlsx"):
    if os.path.exists(archivo):
        libro = load_workbook(archivo)
        hoja = libro.active
    else:
        libro = Workbook()
        hoja = libro.active
        hoja.append(["Nombre", "Puntaje"])

    # si hay datos existentes lleo
    datos = {}
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        nombreExistente, puntajeExistente = fila
        if nombreExistente not in datos or puntajeExistente > datos[nombreExistente]:
            datos[nombreExistente] = puntajeExistente

    # actualiz puntje
    if nombre in datos:
        if puntaje > datos[nombre]:
            datos[nombre] = puntaje
    else:
        datos[nombre] = puntaje

    # si no hay hoja
    libro.remove(hoja)
    hoja = libro.create_sheet(title="Sheet")
    hoja.append(["Nombre", "Puntaje"])
    for nombreGuardado, puntajeGuardado in datos.items():
        hoja.append([nombreGuardado, puntajeGuardado])

    libro.save(archivo)
    print(f"Puntaje guardado: {nombre} - {puntaje}")

def mostrarTop3(archivo="puntajes.xlsx"):
    if not os.path.exists(archivo):
        print("Archivo no encontrado.")
        return

    libro = load_workbook(archivo)
    hoja = libro.active

    datos = []
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        nombre, puntaje = fila
        datos.append((nombre, puntaje))

    datosOrdenados = quicksort(datos)
    top3 = datosOrdenados[:3]

    mensaje = "Top 3 Puntajes:\n"
    for i, (nombre, puntaje) in enumerate(top3, start=1):
        mensaje += f"{i}. {nombre} - {puntaje}\n"

    ventana = tk.Tk()
    ventana.withdraw()
    messagebox.showinfo("Mejores Puntajes", mensaje)
    ventana.destroy()


myApp.mainloop()
