#el coso para el sistema de rankeo
from openpyxl import Workbook, load_workbook
import os

import tkinter as tk
from tkinter import messagebox

def quicksort(lista):
    if len(lista) <= 1:
        return lista
    
    pivote = lista[0]
    iguales = [pivote]
    mayores = []
    menores = []

    for elemento in lista[1:]:
        if elemento[1] > pivote[1]:
            mayores.append(elemento)
        elif elemento[1] < pivote[1]:
            menores.append(elemento)
        else:
            iguales.append(elemento)

    return quicksort(mayores) + iguales + quicksort(menores)

def guardarPuntaje(nombre, puntaje, archivo="puntajes.xlsx"):
    if os.path.exists(archivo):
        libro = load_workbook(archivo)
        hoja = libro.active
    else:
        libro = Workbook()
        hoja = libro.active
        hoja.append(["Nombre", "Puntaje"])

    # si hay datos existentes lleo
    datos = {}
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        nombreExistente, puntajeExistente = fila
        if nombreExistente not in datos or puntajeExistente > datos[nombreExistente]:
            datos[nombreExistente] = puntajeExistente

    # actualiz puntje
    if nombre in datos:
        if puntaje > datos[nombre]:
            datos[nombre] = puntaje
    else:
        datos[nombre] = puntaje

    # si no hay hoja
    libro.remove(hoja)
    hoja = libro.create_sheet(title="Sheet")
    hoja.append(["Nombre", "Puntaje"])
    for nombreGuardado, puntajeGuardado in datos.items():
        hoja.append([nombreGuardado, puntajeGuardado])

    libro.save(archivo)
    print(f"Puntaje guardado: {nombre} - {puntaje}")

#funcion de prueba para mostrar los mejores rankeados
def mostrarTop3(archivo="puntajes.xlsx"):
    if not os.path.exists(archivo):
        print("Archivo no encontrado.")
        return

    libro = load_workbook(archivo)
    hoja = libro.active

    datos = []
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        nombre, puntaje = fila
        datos.append((nombre, puntaje))

    datosOrdenados = quicksort(datos)
    top3 = datosOrdenados[:3]

    mensaje = "Top 3 Puntajes:\n"
    for i, (nombre, puntaje) in enumerate(top3, start=1):
        mensaje += f"{i}. {nombre} - {puntaje}\n"

    ventana = tk.Tk()
    ventana.withdraw()
    messagebox.showinfo("Mejores Puntajes", mensaje)
    ventana.destroy()

# Ejemplo
guardarPuntaje("Yael", 2500)
guardarPuntaje("Cesar", 3200)
guardarPuntaje("Diego", 15000)
guardarPuntaje("Yael", 200000)
mostrarTop3()
