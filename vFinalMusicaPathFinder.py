import tkinter as tk
from tkinter import messagebox
import random
import sys
import time
import pygame #pygame para la musica
from openpyxl import Workbook, load_workbook
import os

# Inicializar pygame.mixer
pygame.mixer.init()

# Cargar y configurar música de fondo
pygame.mixer.music.load("C:/Users/diego/OneDrive/Escritorio/sounds/inGame.mp3")
pygame.mixer.music.set_volume(0.3)
pygame.mixer.music.play(-1)  # Repetir indefinidamente

# Cargar efectos de sonido
sonido_descubrir = pygame.mixer.Sound("C:/Users/diego/OneDrive/Escritorio/sounds/descubrirCasilla.mp3")
sonido_bomba = pygame.mixer.Sound("C:/Users/diego/OneDrive/Escritorio/sounds/explosionBomba.mp3")
sonido_escudo = pygame.mixer.Sound("C:/Users/diego/OneDrive/Escritorio/sounds/escudo.mp3")
sonido_radar = pygame.mixer.Sound("C:/Users/diego/OneDrive/Escritorio/sounds/radar.mp3")
sonido_potenciador = pygame.mixer.Sound("C:/Users/diego/OneDrive/Escritorio/sounds/potenciador.mp3")
sonido_ganar = pygame.mixer.Sound("C:/Users/diego/OneDrive/Escritorio/sounds/victoria.mp3")
sonido_perder = pygame.mixer.Sound("C:/Users/diego/OneDrive/Escritorio/sounds/derrota.mp3")

# Volumen de efectos
for sonido in [sonido_descubrir, sonido_bomba, sonido_escudo, sonido_radar, sonido_potenciador, sonido_ganar, sonido_perder]:
    sonido.set_volume(0.6)

def salir():
    pygame.mixer.music.stop()  # Detener música antes de cerrar
    myApp.destroy()

# Variables globales del minijuego
secuencia = []#Donde se va a guardar el minigame
flechas = {1: "↑", 2: "↓", 3: "←", 4: "→"}#El diccionario para la secuencia
temporizadorPaTodos = None#Es como un ID para el temporizador en caso de que vuelva a caer; se reinicie
tiempo_inicio = None#Para que se vaya actualizando el tiempo conforme pase el tiempo
minaActual=None#Guarda la posicion de la mina presionada
#Las variables de mi tio
myApp = tk.Tk()
myApp.config(bg="black") #Color base de la ventana
myApp.geometry("900x700+300+80") #tamaño establecido y ubicacion inicial


startTime = time.time()
finalTime = 0 # Calculado al final 

gameFrame = tk.Frame(myApp,
                     bg="black",
                    
                     highlightbackground="green yellow",
                    highlightthickness=1)
#gameFrame.pack()

infoFrame = tk.Frame(myApp,
                     highlightbackground="green yellow",
                     highlightthickness=1,
                     bg="black",
                     width=1000
                     )
shieldFrame = tk.Frame(infoFrame,
                       bg="black",
                       highlightbackground="red",
                       highlightthickness=1)
shieldLabel = tk.Label(shieldFrame,
                       text="🛡️0",
                       font = ("",20),
                       bg="black",
                       fg="white")
timeFrame = tk.Frame(infoFrame,
                     bg="black",
                     highlightbackground="red",
                     highlightthickness=1)
timeLabel = tk.Label(timeFrame,
                     bg="black",
                     fg="white",
                     font=("",20))

shieldFrame.grid(row=1,column=1)
shieldLabel.pack(padx=10)
timeFrame.grid(row=1,column=2)
timeLabel.pack(padx=10)


 #-------------------------------------------------
loseFrame = tk.Frame(myApp,
                     width =300,
                     height =200,
                     bg="black")
# Esta es una prueba de la pantalla de derrota
#Libertad de modificarla
# Más detalles están en la función lose(), casi hasta abajo del código

loseLabel = tk.Label(loseFrame, 
                     text = "You Lose",
                     bg="black",
                     highlightbackground="purple",
                     highlightthickness=2,
                     fg="white",
                     font=("arial",50),
                     padx=5,
                     pady=5)

# Frame del minijuego
minijuegoFrame = tk.Frame(myApp,
                          bg="black")

# Variables del minijuego en el frame
muestraFlechas = tk.Label(minijuegoFrame)
info = tk.Label(minijuegoFrame)
etiqueta_tiempo = tk.Label(minijuegoFrame)



#Número de poderes en el tablero
bombs = 3
shields = 4
radars = 3
powerUps = 2
powerMultiplier = 1

remainingShields = 0
usedShields = 0

FinalScore = 0

remainingBombs = bombs
remainingFlags = bombs
currentFlags = 0


map = []
mapGenerated = False
buttons = []
specialButtons = {
    "shield" : 100,
    "radar" : 200,
    "powerUp" : 300,
    "bombs" : 500
}

# Altura del juego
height  = 5 
# Ancho del juego
width = 5

totalCasillas = height * width
casillasCubiertas = 0 # Contador de casillas abiertas

# Configuración de la ventana principal
myApp.title("Mine Sweeper")

# Create a height x width grid for the matrix representation of the game 
# Create a height x width grid for the buttons
for i in range(0,height):
    newList = []
    for j in range(0,width):
        newList.append(0)
    map.append(newList)


for i in range(0,height):
    otherNewList = []
    for j in range(0,width):
        otherNewList.append(0)
    buttons.append(otherNewList)

def actualizarTiempo(): #Actualizar el tiempo para la pantalla del juego
    tiempo_actual = int(time.time() - startTime)
    timeLabel.config(text=f"Tiempo: {tiempo_actual} s")
    timeLabel.after(1000, actualizarTiempo)  # Actualiza cada 1 segundo


def jugar(): #comando para el boton jugar del menu
    global startTime
    startTime = time.time() #El tiempo de inicio inicia en 0
    menu_frame.pack_forget()
    infoFrame.pack()
    gameFrame.pack()
    createButtons() #Crea los botones del buscaminas
    actualizarTiempo()

def menu(): #Funcion para el menu principal
    menu_frame.pack(fill="both", expand="True",)
    name_label.pack(pady=60)
    play_frame.pack(pady=5)
    play_button.pack()
    score_frame.pack()
    score_button.pack()
    
    score_button.config(command = mostrarTop10) #boton de puntuación muestra en una ventana el top 10 puntajes

    
    exit_frame.pack(pady=10)
    exit_button.pack()



#FUNCIONES DEL MINIJUEGO
def actualizarTemporizador():
    global temporizadorPaTodos, tiempo_inicio #Variables que se van a ir actualizando
    tiempoTranscurrido = int(time.time() - tiempo_inicio) #El tiempo de inicio se va a ir actualizando conforme al tiempo
    tiempoRestante = 4 - tiempoTranscurrido #Aquí podríamos añadir mas tiempo; la verdad no se
    etiqueta_tiempo.config(text=f'Tiempo: {tiempoRestante} s',font=("arial",20),fg="white",bg="black") #Esta el la label para el tiempo
    
    if tiempoRestante > 0:
        temporizadorPaTodos = myApp.after(100, actualizarTemporizador) #Se va actualizando el id del temporizador
    else:
        info.config(text="Se acabó el tiempo")#Si se llegó a 0; se acaba el juego
        muestraFlechas.config(text="") #Desaparecen las flechas
        myApp.unbind('<Key>') #Si se modifica; pongan Key con K no k; perdi mas de 2 horas dx
        # Si se acaba el tiempo, va a la pantalla de derrota
        irADerrota()#Se murio la derrota

def iniciarJuego():
    global secuencia, tiempo_inicio, temporizadorPaTodos
    if temporizadorPaTodos:
        myApp.after_cancel(temporizadorPaTodos) #Si ya habia un temporizador; lo elimina y empieza uno nuevo
    tiempo_inicio = time.time() #El tiempo de inicio inicia en 0
    secuencia = [random.randint(1, 4) for i in range(random.randint(5, 10))] #Se va a ingresar a la secuencia 5 a 10 numeros entre 1 y 4
    muestraFlechas.config(text=' '.join([flechas[num] for num in secuencia]),
                          font=("arial",40),
                          fg="green yellow",
                          bg="black",
                          pady=70) #Se muestran las flechas en la etiqueta
    info.config(text="¡Ingresa la secuencia!",font=("arial",20),bg="black",fg="green yellow") #La label informativa 
    actualizarTemporizador() #Se inicia el nuevo temporizador
    myApp.bind('<Key>', teclaPresionada) #Se habilitan las flechas

def teclaPresionada(tecla):
    global secuencia
    convertidor = {'Up': 1, 'Down': 2, 'Left': 3, 'Right': 4} #El diccionario convierte las teclas ingresadas en numeros
    if secuencia: #Mientras la secuencia contenga algo:
        if convertidor[tecla.keysym] == secuencia[0]: #Si el jugador le atinó:
            secuencia.pop(0) #Elimina de la secuencia el numero ingresado
            muestraFlechas.config(text=' '.join([flechas[num] for num in secuencia])) #Se actualiza el visor de flechas
            if not secuencia: #Si acabó satisfactoriamente la secuencia:
                info.config(text="Bomba desarmada") #Muestra la leyenda
                myApp.unbind('<Key>') #Se quita lo de las key
                if temporizadorPaTodos:
                    myApp.after_cancel(temporizadorPaTodos) #Detiene y cancela el temporizador
                # Si gana, regresa al juego principal
                volverAlJuego() #Se continua en la partida
        else:
            info.config(text="La bomba explotó") #Si falla; pierde la partida
            muestraFlechas.config(text="") #Ya no se muestra la secuencia en el label
            myApp.unbind('<Key>') #Se quitan los permisos para las flechitas
            if temporizadorPaTodos:
                myApp.after_cancel(temporizadorPaTodos) #Se detiene el temporizador
            # Si explota la bomba, va a la pantalla de derrota
            irADerrota() #Va a la pantalla de derrota

def mostrarMinijuego():
    gameFrame.pack_forget() #Se oculta el widget del buscaminas
    minijuegoFrame.pack()  #Se muestra el minijuego
    muestraFlechas.pack() #Se muestra la secuencia
    info.pack(pady=40) #Se muestra la info
    etiqueta_tiempo.pack() #Se muestra el tiempo
    iniciarJuego() #Se inicia el juego

def volverAlJuego():
    minijuegoFrame.pack_forget() #Se oculta el minijuego si se gana
    muestraFlechas.config(text="") #Se ocultan las labels
    info.config(text="") 
    etiqueta_tiempo.config(text="")
    gameFrame.pack() #Se vuelve al buscaminas
    global mina_actual
    if mina_actual is not None:
        h, w = mina_actual
        map[h][w] = -1
        buttons[h][w]["text"] = "🪛"
        mina_actual = None

def irADerrota():
    minijuegoFrame.pack_forget() #Se oculta el minijuego
    lose() #Se pierde

def rePlay():
    global map, buttons, mapGenerated, remainingShields, startTime, casillasCubiertas, currentFlags, remainingBombs, remainingFlags
    # Oculta el frame de derrota y muestra el del juego
    startTime = time.time()
    loseFrame.pack_forget()
    infoFrame.pack()
    gameFrame.pack()
    # Limpiar el frame del juego
    for widget in gameFrame.winfo_children():
        widget.destroy()
    # Reiniciar variables
    map = []
    buttons = []
    mapGenerated = False
    remainingShields = 0
    casillasCubiertas = 0
    currentFlags = 0
    remainingBombs = bombs
    remainingFlags = bombs
    # Volver a crear la matriz y los botones
    for i in range(height):
        newList = []
        for j in range(width):
            newList.append(0)
        map.append(newList)
    for i in range(height):
        otherNewList = []
        for j in range(width):
            otherNewList.append(0)
        buttons.append(otherNewList)
    for i in range(height):
        for j in range(width):
            currentName = str(i)+str(False)+str(j)
            cell_frame = tk.Frame(gameFrame,
                                  highlightbackground="purple",
                                  highlightthickness=1,
                                  bg="black",
                                  width=40,
                                  height=40)
            cell_frame.grid(row=i, column=j)
            
            button  = tk.Button(cell_frame,
                    name = str(i)  +str(j),
                    bg="black",
                    
                    border=0,
                    fg="white",
                    width=4,
                    height=2,
                    activebackground="green yellow",
                    command =  lambda buttonName = currentName: buttonClick(map,buttonName,buttons))
            button.pack()
            if sys.platform == "darwin":
                button.bind("<Button-2>", flag)
            else:
                button.bind("<Button-3>", flag)
            buttons[i][j] = button

#FUNCIONES DEL BUSCAMINAS
#Función para visualizar el mapa en la terminal
#Solo usar para depuración 
def printMap(gameMap: list):
    for i in range(0,len(gameMap)):
        print(gameMap[i])

# Adding every game object:
def addGameObject(gameMap: list, x_avoid :int, y_avoid : int, nProperty: int, propertyName :str):
    # Making sure the first cell is right next to a bomb
    startingX = x_avoid
    startingY = y_avoid
    if propertyName == "bombs":
        if startingX == width - 1:
            startingX = x_avoid -1
        if startingY == height - 1:
            startingY = y_avoid - 1
        if x_avoid != width - 1 and y_avoid != height - 1:
            gameMap[startingY+1][startingX+1] = specialButtons[propertyName]
        else:
            gameMap[startingY][startingX] = specialButtons[propertyName]
        nProperty -= 1
            
    nProperty -= 1
    while nProperty > 0:
        for i in range(0,height):
            for j in range(0,width):
                if i!= y_avoid and j!=x_avoid and gameMap[i][j] == 0:
                    if random.randint(0,100)<5 and nProperty > 0:
                        gameMap[i][j] = specialButtons[propertyName]
                        nProperty -= 1
                        if propertyName == "powerUp":
                            gameMap[i][j] = specialButtons[propertyName] + random.randint(15,50)

# Initialize the map with 0s and place all objects randomly
def initializeMap(gameMap: list, x_0 : int, y_0:int, nBombs: int, nShields: int, nRadars:int, nPowerUps: int):
    global mapGenerated
    addGameObject(gameMap, x_0, y_0, nBombs, "bombs")
    addGameObject(gameMap, x_0, y_0, nShields, "shield")
    addGameObject(gameMap, x_0, y_0, nRadars, "radar")
    addGameObject(gameMap, x_0, y_0, nPowerUps, "powerUp")
    """
    # Fills the map with number of surrounding bombs
    for i in range(0,height):
            for j in range(0,width):
                if map[i][j] != 500:
                    map[i][j] = getNeighborBombs(map,j,i)
    """
    mapGenerated = True

# Gets the count of bombs surrounding a cell
def getNeighborBombs(gameMap: list, x : int, y : int) -> int:
    count = 0
    for i in range(-1, 2):
        for j in range(-1, 2):
            if 0 <= x + i < width and 0 <= y + j < height:
                if gameMap[y + j][x + i] == 500 or gameMap[y + j][x + i] == -1 :
                    count += 1
    return count

def radarEffect(gameMap: list, x: int, y: int, buttons: list): 
    # This function will reveal all the cells in a 3x3 area around the radar
    global casillasCubiertas
    global powerMultiplier
    global remainingShields
    unclickedText = ["🚩", ""]  # Text for unclicked cells
    for i in range(-1, 2):
        for j in range(-1, 2):
            if 0 <= x + i < width and 0 <= y + j < height:
                
                if i!= 0 or j!=0 :
                    
                    if buttons[y + j][x + i]["text"]  in unclickedText:
                        casillasCubiertas +=1
                        print(f"Casillas cubiertas added at {x+i}, {y+j}: {casillasCubiertas}")
                        if gameMap[y + j][x + i] == 0:
                            buttons[y + j][x + i]["text"] = getNeighborBombs(gameMap, x + i, y + j)
                        elif gameMap[y + j][x + i]  ==100:
                            shieldEffect(gameMap, x + i, y + j, buttons)
                            buttons[y + j][x + i]["text"] = str('🛡️')
                        elif gameMap[y + j][x + i] == 200:
                            buttons[y + j][x + i]["text"] = str('📡')
                        elif gameMap[y + j][x + i] >300 and gameMap[y + j][x + i] < 500:
                            powerUpEffect(gameMap, x + i, y + j, buttons)
                        # if it is a bomb
                        elif gameMap[y + j][x + i] == 500 :
                            buttons[y + j][x + i]["text"] = str('💣')
                    sonido_radar.play()  # Efecto de radar

def shieldEffect( gameMap: list, x: int, y: int, buttons: list):
    # Adds a shield to the game
    global remainingShields
    if buttons[y][x]["text"] == "🚩" or buttons[y][x]["text"] == "":
        remainingShields += 1
        shieldLabel.config(text="🛡️" + str(remainingShields))
        #buttons[y][x]["text"] = str('\t🛡️\n'+ str(getNeighborBombs(gameMap,x,y)))
        buttons[y][x]["text"] = "🛡️"
        sonido_escudo.play()  # Efecto al obtener escudo



def lose(lost: bool = True):
    global finalTime
    global casillasCubiertas
    global totalCasillas
    #print(finalTime)
    finalTime = time.time() - startTime
    FinalScore = (finalTime*-0.001 + 100)* powerMultiplier * ((casillasCubiertas + currentFlags) / totalCasillas)
    FinalScore = round(FinalScore, 2) * 10000
    print(f"Final Score: {FinalScore}") # Debugging line to check the final score

    gameFrame.pack_forget()
    infoFrame.pack_forget()

    for widget in loseFrame.winfo_children():
        widget.destroy()
    loseFrame.pack()
    # Crea el label de "You Lose" aquí
    loseLabel = tk.Label(loseFrame,
                         text="You Lose",
                         bg="black",
                         highlightbackground="purple",
                         highlightthickness=2,
                         fg="purple",
                         font=("arial", 50),
                         padx=5,
                         pady=5)
    sonido_perder.play()#musica si pierdes
    if not lost:
        loseLabel.config(text="You Win")
        sonido_ganar.play()#musica si ganas

    losetime = tk.Label(loseFrame,
                        font=("arial", 20),
                        text="Tiempo:\t" + str(round(finalTime, 2)),
                        bg="black",
                        fg="white")

    loseScore = tk.Label(loseFrame,
                         font=("arial", 20),
                         text="Puntuacion:\t" + str(round(FinalScore, 2)),
                         bg="black",
                         fg="white")

    lose_rePlay_Frame = tk.Frame(loseFrame,
                                highlightbackground="green yellow",
                                highlightthickness=1,
                                bg="black")
    lose_rePlay_Button = tk.Button(lose_rePlay_Frame,
                                  text="Reintentarlo",
                                  fg="green yellow",
                                  bg="black",
                                  font=("arial", 20),
                                  command=rePlay,
                                  border=0)
    lose_Menu_Frame = tk.Frame(loseFrame,
                             highlightbackground="green yellow",
                             highlightthickness=1,
                             bg="black")
    lose_Menu_Button = tk.Button(lose_Menu_Frame,
                                text="Menu",
                                fg="green yellow",
                                bg="black",
                                font=("arial", 20),
                                command=lose_menu,
                                border=0,
                                activebackground="green yellow")
    lose_exit_Frame = tk.Frame(loseFrame,
                             highlightbackground="magenta2",
                             highlightthickness=1,
                             bg="black")
    lose_exit_Button = tk.Button(lose_exit_Frame,
                                text="Salir",
                                fg="magenta2",
                                bg="black",
                                font=("arial", 20),
                                command=salir,
                                border=0,
                                activebackground="magenta2")
    

    # pedir nombre y guardar puntaje
    def guardar():
        nombre = nombreEntry.get()
        if nombre.strip():
            guardarPuntaje(nombre, FinalScore)
        else:
            messagebox.showwarning("Aviso", "Por favor, ingresa un nombre.")

    formFrame = tk.Frame(loseFrame, bg="black")
    nombreLabel = tk.Label(formFrame,
                           text="Nombre:",
                           bg="black",
                           fg="white",
                           font=("arial", 14))
    nombreEntry = tk.Entry(formFrame,
                           font=("arial", 14),
                           bg="gray20",
                           fg="white",
                           insertbackground="white",
                           relief="flat",
                           width=15)
    guardarBtn = tk.Button(formFrame,
                          text="Guardar",
                          command=guardar,
                          font=("arial", 14),
                          bg="green yellow",
                          fg="black",
                          activebackground="yellow",
                          activeforeground="black",
                          relief="flat",
                          padx=8,
                          pady=3)

    # Empaquetar la pedida de nombre en horizontal
    nombreLabel.pack(side="left", padx=(0,5))
    nombreEntry.pack(side="left", padx=(0,5))
    guardarBtn.pack(side="left")

    # Empaqueta los widgets en el orden correcto
    loseLabel.pack(pady=50)
    losetime.pack()
    loseScore.pack(pady=20)
    lose_rePlay_Button.pack()
    lose_rePlay_Frame.pack(pady=40)
    lose_Menu_Button.pack()
    lose_Menu_Frame.pack(pady=5)
    lose_exit_Button.pack()
    lose_exit_Frame.pack(pady=20)
    formFrame.pack(pady=15) # frame para agrupar y alinear el formulario de nombre y botón guarda

def lose_menu():
    loseFrame.pack_forget()
    menu_frame.pack()

def powerUpEffect(gameMap: list, x: int, y: int, buttons: list):
    global powerMultiplier
    if buttons[y][x]["text"] == "" or buttons[y][x]["text"] == "🚩":
        #buttons[y][x]["text"] = str(['⚡', getNeighborBombs(gameMap,x,y)])
        #buttons[y][x]["text"] = str(['⚡', getNeighborBombs(gameMap,x,y)])
        buttons[y][x]["text"] = str('⚡')
        buttons[y][x]["text"] = str('⚡')
        powerMultiplier = powerMultiplier * (1+((gameMap[y][x] - specialButtons["powerUp"])/100))
        print(f"Power multiplier: {powerMultiplier}")  # Debugging line to check the multiplier value
        sonido_potenciador.play() #sonido para el potenciador
    

# Defines behavior for left click on a button
def buttonClick(gameMap: list, name : str,buttonSet : list):
    global mapGenerated
    global remainingShields
    global remainingFlags
    global remainingBombs
    global totalCasillas
    global casillasCubiertas
    global currentFlags
    global remainingBombs
    unclickedText = ["🚩", ""]  # Text for unclicked cell   s
    

    usedShields=0
    buttonHeight =int(name[:name.index('F')])
    buttonWidth = int(name[name.index('e')+1:])

    if buttons[buttonHeight][buttonWidth]["text"] in unclickedText and gameMap[buttonHeight][buttonWidth] != 500:
        casillasCubiertas +=1

    print(f"Clicked squares: {casillasCubiertas}, flags: {currentFlags}, total: {totalCasillas}")
    if casillasCubiertas == totalCasillas or casillasCubiertas + currentFlags == totalCasillas:
        lose(False)

    

    if not mapGenerated:
        initializeMap(gameMap, buttonWidth, buttonHeight, bombs, shields, radars, powerUps)
        printMap(gameMap)
    
    if gameMap[buttonHeight][buttonWidth] == -1:
        return

    if buttonSet[buttonHeight][buttonWidth]["text"] == "🚩":
        remainingFlags += 1

    if buttonSet[buttonHeight][buttonWidth]["text"]  in unclickedText:
        
        if gameMap[buttonHeight][buttonWidth] == 500:
            
            if buttonSet[buttonHeight][buttonWidth]["text"] == "🚩":
                remainingFlags -= 1
            if remainingShields == 0:
                global mina_actual
                mina_actual = (buttonHeight, buttonWidth)
                mostrarMinijuego()
                
                
            else:
                remainingShields-=1
                shieldLabel.config(text="🛡️"+str(remainingShields))
                usedShields +=1
            buttonSet[buttonHeight][buttonWidth]["text"] = str('💣')
            sonido_bomba.play()  # Efecto bomba si no hay escudo
            remainingBombs -= 1
            casillasCubiertas +=1
            
        elif gameMap[buttonHeight][buttonWidth] == 100:
            shieldEffect(gameMap, buttonWidth, buttonHeight, buttonSet)
            
        elif gameMap[buttonHeight][buttonWidth] == 200:
            radarEffect(gameMap, buttonWidth, buttonHeight, buttonSet)
            buttons[buttonHeight][buttonWidth]["text"] = str('📡')
        elif gameMap[buttonHeight][buttonWidth] >= 300 and gameMap[buttonHeight][buttonWidth] < 500: 
            powerUpEffect(gameMap, buttonWidth, buttonHeight, buttonSet)
        else:
            buttonSet[buttonHeight][buttonWidth]["text"] =getNeighborBombs(gameMap,buttonWidth, buttonHeight)
            sonido_descubrir.play() #SONIDO AL DESCUBRIR CASILLA
    if casillasCubiertas == totalCasillas or casillasCubiertas + currentFlags == totalCasillas:
        lose(False) 

# Defines flagging a button
def flag( event ):
    global remainingFlags
    global currentFlags
    if event.widget["text"] == "🚩":
        remainingFlags +=1
        
        event.widget.config(text="")
    elif event.widget["text"] == "":
        remainingFlags -= 1
        currentFlags +=1
        event.widget.config(text="🚩")
    if casillasCubiertas == totalCasillas or casillasCubiertas + currentFlags == totalCasillas:
        lose(False)


if sys.platform == "darwin":
    for i in range(height):
        gameFrame.rowconfigure(i, weight=1)
    for i in range(width):
        gameFrame.columnconfigure(i, weight=1)

# Create buttons for the grid
def createButtons():
    global buttons
    global map
    for i in range(0,height):
        for j in range(0,width):
            currentName = str(i)+str(False)+str(j)
            cell_frame = tk.Frame(gameFrame,
                                highlightbackground="purple",
                                highlightthickness=1,
                                bg="black",
                                width=40,
                                height=40)
            button  =tk.Button(cell_frame,
                    name = str(i)  +str(j),
                    bg="black",
                    border=0,
                    font=("",20),
                    fg="white",
                    width=4,
                    height=2,
                    activebackground="green yellow",
                    command =  lambda buttonName = currentName: buttonClick(map,buttonName,buttons))
            #print(str(i)+str(j))
            cell_frame.pack_propagate(False) 
            cell_frame.grid(
                row = i,
                column = j
            )
            
            '''
            button.grid(
                row = i,
                column = j
            )
            '''
            button.pack()
            #checks os for right click
            if sys.platform == "darwin":
                button.bind("<Button-2>", flag)
            else:
                button.bind("<Button-3>", flag)
            #print(type(map[i][j]))
            buttons[i][j] = button

#Variable para el menu:
menu_frame = tk.Frame(myApp,bg="black")
name_label = tk.Label(menu_frame,
                          bg="black",
                          text="Pathfinder",
                          fg="white",
                          font=("arial",20),
                          anchor='center',
                          highlightbackground="purple",
                          highlightthickness=3,
                          pady=5)
play_frame = tk.Frame(menu_frame,
                          highlightbackground="green yellow",
                          highlightthickness=1)
play_button = tk.Button(play_frame,
                            width="15",
                            text="Iniciar juego",
                            bg="black",
                            fg="white",
                            font=("arial",15),
                            pady=5,
                            border=0,
                            activebackground="green yellow",
                            command=jugar)
score_frame = tk.Frame(menu_frame,
                           highlightbackground="green yellow",
                           highlightthickness=1)
score_button = tk.Button(score_frame,
                            width="15",
                            text="Puntuacion",
                            bg="black",
                            fg="white",
                            font=("arial",15),
                            pady=5,
                            border=0,
                            activebackground="green yellow")
exit_frame = tk.Frame(menu_frame,
                            highlightbackground="magenta2",
                            highlightthickness=1)
exit_button = tk.Button(exit_frame,
                            width="5",
                            text="Salir",
                            bg="black",
                            fg="white",
                            font=("arial",10),
                            pady=5,
                            command=salir,
                            border=0,
                            activebackground="magenta2")

def quicksort(lista):
    if len(lista) <= 1:
        return lista
    
    pivote = lista[0]
    iguales = [pivote]
    mayores = []
    menores = []

    for elemento in lista[1:]:
        if elemento[1] > pivote[1]:
            mayores.append(elemento)
        elif elemento[1] < pivote[1]:
            menores.append(elemento)
        else:
            iguales.append(elemento)

    return quicksort(mayores) + iguales + quicksort(menores)

def guardarPuntaje(nombre, puntaje, archivo="puntajes.xlsx"):
    if os.path.exists(archivo):
        libro = load_workbook(archivo)
        hoja = libro.active
    else:
        libro = Workbook()
        hoja = libro.active
        hoja.append(["Nombre", "Puntaje"])

    # si hay datos existentes lleo
    datos = {}
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        nombreExistente, puntajeExistente = fila
        if nombreExistente not in datos or puntajeExistente > datos[nombreExistente]:
            datos[nombreExistente] = puntajeExistente

    # actualiz puntje
    if nombre in datos:
        if puntaje > datos[nombre]:
            datos[nombre] = puntaje
    else:
        datos[nombre] = puntaje

    # si no hay hoja
    libro.remove(hoja)
    hoja = libro.create_sheet(title="Sheet")
    hoja.append(["Nombre", "Puntaje"])
    for nombreGuardado, puntajeGuardado in datos.items():
        hoja.append([nombreGuardado, puntajeGuardado])

    libro.save(archivo)
    print(f"Puntaje guardado: {nombre} - {puntaje}")


def mostrarTop10(archivo="puntajes.xlsx"):
    if not os.path.exists(archivo):
        print("Archivo no encontrado.")
        return

    libro = load_workbook(archivo)
    hoja = libro.active

    datos = []
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        nombre, puntaje = fila
        datos.append((nombre, puntaje))

    datosOrdenados = quicksort(datos)
    top10 = datosOrdenados[:10]

    mensaje = ""
    for i, (nombre, puntaje) in enumerate(top10, start=1):
        mensaje += f"{i}. {nombre} - {puntaje}\n"

    '''
    ventana = tk.Tk()
    ventana.withdraw()
    messagebox.showinfo("Mejores Puntajes", mensaje)
    ventana.destroy()
    '''

    menu_frame.pack_forget()
    top_Frame.pack()
    top_Label.config(text=mensaje)
    return_Button_Frame.pack()

def return_Frame():
    top_Frame.pack_forget()
    return_Button_Frame.pack_forget()
    menu_frame.pack()

top_Frame = tk.Frame(myApp,
                     bg="black",
                     highlightbackground="magenta3",
                     highlightthickness=1,
                     width="500",
                     height="600")
top_Frame.pack_propagate(False)
top_title_Label = tk.Label(top_Frame,
                           bg="black",
                           font = ("",30),
                           fg="white",
                           text="Top Score")
top_Label = tk.Label(top_Frame,
                     text="",
                     bg="black",
                     fg="white",
                     font = ("",20),
                     pady=20)
return_Button_Frame = tk.Frame(myApp,
                               bg="black",
                               highlightbackground="green yellow",
                               highlightthickness=1)
return_Button = tk.Button(return_Button_Frame,
                          bg="black",
                          fg="white",
                          text="Regresar",
                          font = ("",20),
                          command=return_Frame
                          )
top_title_Label.pack(pady=50)
top_Label.pack()
return_Button.pack()

#Inicia el juego
menu()
myApp.mainloop()
